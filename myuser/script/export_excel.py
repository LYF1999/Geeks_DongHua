#  coding=utf-8
from __future__ import unicode_literals
from openpyxl import Workbook
from myuser.models import User


def export_member():
    wb = Workbook()
    new_ws = wb.create_sheet(title='students')
    row = 1

    users = User.objects.filter(is_member=True)

    for user in users:
        new_ws.cell(row=row, column=1).value = row
        new_ws.cell(row=row, column=2).value = user.name
        row += 1

    wb.save(filename='student.xlsx')
