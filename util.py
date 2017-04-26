#  coding=utf-8

from openpyxl import Workbook

import re
import pytz


def is_valid_tel(tel):
    if not re.match(r'1\d{10}', tel):
        return False
    return True


def format_time(time):
    return time.astimezone(tz=pytz.timezone('Asia/Shanghai')).strftime('%Y年%m月%d日 %H:%M:%S').decode('utf8')


def get_fix_excel():
    from fix.models import Fix
    fixs = Fix.objects.filter(is_fix=True).all().order_by('-appointment_time')[:70]
    index = 1
    row = 1
    wb = Workbook()
    new_ws = wb.create_sheet(title='students')
    for fix in fixs:
        new_ws.cell(row=row, column=1).value = index
        new_ws.cell(row=row, column=2).value = '上海'
        new_ws.cell(row=row, column=3).value = '东华大学'
        new_ws.cell(row=row, column=4).value = fix.name
        new_ws.cell(row=row, column=5).value = fix.tel
        new_ws.cell(row=row, column=6).value = '2016'
        new_ws.cell(row=row, column=7).value = fix.model
        new_ws.cell(row=row, column=8).value = fix.fault.name
        new_ws.cell(row=row, column=9).value = fix.appointment_time
        row += 1
        index += 1

    wb.save(filename='fix.xlsx')
